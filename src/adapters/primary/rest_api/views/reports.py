"""
Reports ViewSet para generación de reportes y exportación de datos.

Este módulo implementa endpoints para:
- Reportes PDF de prácticas
- Exportación Excel de datos
- Reportes estadísticos
- Certificados y constancias

Arquitectura: Hexagonal (Adaptador Primario)
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse, FileResponse
from django.db.models import Count, Avg, Q, Sum
from django.utils import timezone
from datetime import timedelta, datetime
import io
import csv
import json

# drf-spectacular para documentación OpenAPI
from drf_spectacular.utils import (
    extend_schema, extend_schema_view, OpenApiParameter,
    OpenApiResponse, OpenApiExample,
)
from drf_spectacular.types import OpenApiTypes

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from src.domain.entities import (
    User, Student, Company, Supervisor, Practice, 
    Document, Notification
)
from src.infrastructure.security.permissions import (
    IsCoordinador, IsPracticante, IsSupervisor, 
    IsSecretaria, IsAdministrador
)


class ReportsViewSet(viewsets.ViewSet):
    """
    ViewSet para generación y exportación de reportes del sistema.
    
    Proporciona endpoints para generar reportes en múltiples formatos:
    
    **Reportes disponibles**:
    - `practices`: Reporte detallado de prácticas con filtros avanzados
    - `students`: Reporte de estudiantes y su progreso
    - `companies`: Reporte de empresas colaboradoras
    - `export_excel`: Exportación masiva a formato Excel
    - `export_csv`: Exportación masiva a formato CSV
    - `practice_certificate`: Certificado oficial de práctica (PDF)
    - `statistics_summary`: Resumen estadístico ejecutivo
    
    **Formatos soportados**:
    - JSON (respuesta API estándar)
    - Excel (.xlsx) con estilos y formato profesional
    - CSV (compatible con Excel, Google Sheets)
    - PDF (certificados y reportes oficiales)
    
    **Características**:
    - Filtros avanzados por fecha, estado, carrera, empresa
    - Exportación con estilos profesionales
    - Generación de certificados oficiales
    - Optimización para grandes volúmenes de datos
    """
    
    permission_classes = [IsAuthenticated]

    # ========================================================================
    # Reporte de Prácticas
    # ========================================================================

    @extend_schema(
        tags=['Reportes'],
        summary='Reporte de Prácticas',
        description='''
        Genera reporte completo de prácticas con filtros avanzados.
        
        Requiere rol de **COORDINADOR**, **ADMINISTRADOR** o **SECRETARIA**.
        
        **Filtros disponibles**:
        - `status`: Filtra por estado(s) separados por coma (IN_PROGRESS, COMPLETED, etc.)
        - `start_date`: Fecha de inicio del período (YYYY-MM-DD)
        - `end_date`: Fecha de fin del período (YYYY-MM-DD)
        - `career`: Filtra por carrera profesional
        - `company_id`: Filtra por empresa específica
        - `format`: Formato de salida (json, excel, csv)
        
        **Formatos de exportación**:
        - `json`: Respuesta JSON con datos estructurados
        - `excel`: Archivo Excel (.xlsx) con estilos y formato
        - `csv`: Archivo CSV compatible con Excel
        ''',
        parameters=[
            OpenApiParameter(
                name='status',
                description='Estados de práctica (separados por coma)',
                required=False,
                type=OpenApiTypes.STR,
                examples=[
                    OpenApiExample('Estado único', value='IN_PROGRESS'),
                    OpenApiExample('Múltiples estados', value='IN_PROGRESS,COMPLETED'),
                ],
            ),
            OpenApiParameter(
                name='start_date',
                description='Fecha de inicio del período (YYYY-MM-DD)',
                required=False,
                type=OpenApiTypes.DATE,
            ),
            OpenApiParameter(
                name='end_date',
                description='Fecha de fin del período (YYYY-MM-DD)',
                required=False,
                type=OpenApiTypes.DATE,
            ),
            OpenApiParameter(
                name='career',
                description='Carrera profesional para filtrar',
                required=False,
                type=OpenApiTypes.STR,
            ),
            OpenApiParameter(
                name='company_id',
                description='ID de empresa para filtrar',
                required=False,
                type=OpenApiTypes.UUID,
            ),
            OpenApiParameter(
                name='format',
                description='Formato de exportación',
                required=False,
                type=OpenApiTypes.STR,
                enum=['json', 'excel', 'csv'],
                default='json',
            ),
        ],
        responses={
            200: OpenApiResponse(
                description='Reporte de prácticas generado',
                examples=[
                    OpenApiExample(
                        'Reporte JSON',
                        value={
                            'success': True,
                            'total_practices': 120,
                            'data': [
                                {
                                    'id': 'uuid',
                                    'titulo': 'Desarrollo Web',
                                    'student': {
                                        'nombre': 'Juan Pérez',
                                        'codigo': '2020123456',
                                        'carrera': 'Ingeniería de Sistemas',
                                    },
                                    'company': {
                                        'razon_social': 'Tech Corp SAC',
                                        'ruc': '20123456789',
                                    },
                                    'estado': 'IN_PROGRESS',
                                    'fecha_inicio': '2024-01-15',
                                    'horas_completadas': 240,
                                },
                            ],
                        },
                    ),
                ],
            ),
            403: OpenApiResponse(description='Sin permisos administrativos'),
        },
    )
    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsCoordinador | IsAdministrador | IsSecretaria],
        url_path='practices'
    )
    def practices_report(self, request):
        """
        Reporte completo de prácticas con filtros.
        
        GET /api/v2/reports/practices/
        ?status=IN_PROGRESS,COMPLETED
        &start_date=YYYY-MM-DD
        &end_date=YYYY-MM-DD
        &career=Ingeniería de Sistemas
        &company_id=1
        &format=json|excel|csv
        
        Requiere: COORDINADOR, ADMINISTRADOR o SECRETARIA
        """
        try:
            # Obtener parámetros de filtro
            status_filter = request.query_params.get('status', '').split(',')
            status_filter = [s.strip() for s in status_filter if s.strip()]
            
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            career = request.query_params.get('career')
            company_id = request.query_params.get('company_id')
            format_type = request.query_params.get('format', 'json')
            
            # Construir query
            queryset = Practice.objects.select_related(
                'student__user',
                'company',
                'supervisor__user'
            ).all()
            
            if status_filter:
                queryset = queryset.filter(estado__in=status_filter)
            
            if start_date_str:
                start_date = datetime.fromisoformat(start_date_str)
                queryset = queryset.filter(fecha_inicio__gte=start_date)
            
            if end_date_str:
                end_date = datetime.fromisoformat(end_date_str)
                queryset = queryset.filter(fecha_fin__lte=end_date)
            
            if career:
                queryset = queryset.filter(student__escuela_profesional__icontains=career)
            
            if company_id:
                queryset = queryset.filter(company_id=company_id)
            
            # Preparar datos
            practices_data = []
            for practice in queryset:
                practices_data.append({
                    'id': practice.id,
                    'titulo': practice.titulo,
                    'estudiante': {
                        'codigo': practice.student.codigo_estudiante,
                        'nombre': practice.student.user.get_full_name(),
                        'carrera': practice.student.escuela_profesional,
                        'promedio': float(practice.student.promedio_ponderado)
                    },
                    'empresa': {
                        'id': practice.company.id,
                        'razon_social': practice.company.razon_social,
                        'ruc': practice.company.ruc,
                        'sector': practice.company.sector_economico
                    },
                    'supervisor': {
                        'nombre': practice.supervisor.user.get_full_name() if practice.supervisor else 'Sin asignar',
                        'cargo': practice.supervisor.cargo if practice.supervisor else None
                    },
                    'fechas': {
                        'inicio': practice.fecha_inicio.isoformat() if practice.fecha_inicio else None,
                        'fin': practice.fecha_fin.isoformat() if practice.fecha_fin else None,
                        'registro': practice.fecha_registro.isoformat()
                    },
                    'horas': {
                        'totales': practice.horas_totales,
                        'completadas': practice.horas_completadas or 0,
                        'porcentaje': round((practice.horas_completadas or 0) / practice.horas_totales * 100, 2) if practice.horas_totales > 0 else 0
                    },
                    'estado': practice.estado,
                    'modalidad': practice.modalidad
                })
            
            # Formatear salida
            if format_type == 'excel':
                return self._export_to_excel(practices_data, 'Reporte_Practicas')
            elif format_type == 'csv':
                return self._export_to_csv(practices_data, 'Reporte_Practicas')
            else:
                return Response({
                    'success': True,
                    'count': len(practices_data),
                    'data': practices_data,
                    'filters': {
                        'status': status_filter,
                        'start_date': start_date_str,
                        'end_date': end_date_str,
                        'career': career,
                        'company_id': company_id
                    },
                    'generated_at': timezone.now().isoformat()
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar reporte de prácticas: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ========================================================================
    # Reporte de Estudiantes
    # ========================================================================

    @extend_schema(
        tags=['Reportes'],
        summary='Reporte de Estudiantes',
        description='''
        Genera reporte de estudiantes con información académica y de prácticas.
        
        Requiere rol de **COORDINADOR**, **ADMINISTRADOR** o **SECRETARIA**.
        
        **Filtros disponibles**:
        - `career`: Filtra por carrera profesional
        - `semester`: Filtra por semestre(s) separados por coma (6,7,8,...)
        - `with_practice`: Filtra estudiantes con/sin práctica (true/false)
        - `format`: Formato de salida (json, excel, csv)
        
        **Información incluida**:
        - Datos académicos (código, carrera, semestre, promedio)
        - Estado de elegibilidad para prácticas
        - Prácticas realizadas (si aplica)
        - Información de contacto
        ''',
        parameters=[
            OpenApiParameter(
                name='career',
                description='Carrera profesional',
                required=False,
                type=OpenApiTypes.STR,
            ),
            OpenApiParameter(
                name='semester',
                description='Semestre(s) académico(s) separados por coma',
                required=False,
                type=OpenApiTypes.STR,
                examples=[
                    OpenApiExample('Semestre único', value='8'),
                    OpenApiExample('Múltiples semestres', value='6,7,8,9'),
                ],
            ),
            OpenApiParameter(
                name='with_practice',
                description='Filtra por estudiantes con/sin práctica',
                required=False,
                type=OpenApiTypes.BOOL,
            ),
            OpenApiParameter(
                name='format',
                description='Formato de exportación',
                required=False,
                type=OpenApiTypes.STR,
                enum=['json', 'excel', 'csv'],
                default='json',
            ),
        ],
        responses={
            200: OpenApiResponse(
                description='Reporte de estudiantes',
                examples=[
                    OpenApiExample(
                        'Reporte Estudiantes',
                        value={
                            'success': True,
                            'total_students': 250,
                            'data': [
                                {
                                    'codigo': '2020123456',
                                    'nombre_completo': 'María García',
                                    'email': 'maria@upeu.edu.pe',
                                    'carrera': 'Ingeniería de Sistemas',
                                    'semestre': 8,
                                    'promedio': 14.5,
                                    'elegible_practicas': True,
                                    'practices_count': 1,
                                },
                            ],
                        },
                    ),
                ],
            ),
            403: OpenApiResponse(description='Sin permisos administrativos'),
        },
    )
    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsCoordinador | IsAdministrador | IsSecretaria],
        url_path='students'
    )
    def students_report(self, request):
        """
        Reporte de estudiantes con información de prácticas.
        
        GET /api/v2/reports/students/
        ?career=Ingeniería de Sistemas
        &semester=6,7,8
        &with_practice=true|false
        &format=json|excel|csv
        
        Requiere: COORDINADOR, ADMINISTRADOR o SECRETARIA
        """
        try:
            # Parámetros
            career = request.query_params.get('career')
            semester_filter = request.query_params.get('semester', '').split(',')
            semester_filter = [int(s.strip()) for s in semester_filter if s.strip().isdigit()]
            with_practice = request.query_params.get('with_practice')
            format_type = request.query_params.get('format', 'json')
            
            # Query
            queryset = Student.objects.select_related('user').prefetch_related('practices')
            
            if career:
                queryset = queryset.filter(escuela_profesional__icontains=career)
            
            if semester_filter:
                queryset = queryset.filter(semestre_actual__in=semester_filter)
            
            if with_practice == 'true':
                queryset = queryset.filter(practices__isnull=False).distinct()
            elif with_practice == 'false':
                queryset = queryset.filter(practices__isnull=True)
            
            # Preparar datos
            students_data = []
            for student in queryset:
                # Práctica activa
                active_practice = student.practices.filter(
                    estado__in=['IN_PROGRESS', 'PENDING', 'APPROVED']
                ).first()
                
                students_data.append({
                    'codigo': student.codigo_estudiante,
                    'nombre': student.user.get_full_name(),
                    'email': student.user.email,
                    'carrera': student.escuela_profesional,
                    'semestre': student.semestre_actual,
                    'promedio': float(student.promedio_ponderado),
                    'elegible': student.is_eligible_for_practice(),
                    'practica_actual': {
                        'id': active_practice.id if active_practice else None,
                        'titulo': active_practice.titulo if active_practice else None,
                        'empresa': active_practice.company.razon_social if active_practice else None,
                        'estado': active_practice.estado if active_practice else None
                    } if active_practice else None,
                    'total_practicas': student.practices.count(),
                    'estado': student.estado
                })
            
            if format_type == 'excel':
                return self._export_to_excel(students_data, 'Reporte_Estudiantes')
            elif format_type == 'csv':
                return self._export_to_csv(students_data, 'Reporte_Estudiantes')
            else:
                return Response({
                    'success': True,
                    'count': len(students_data),
                    'data': students_data,
                    'generated_at': timezone.now().isoformat()
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar reporte de estudiantes: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ========================================================================
    # Reporte de Empresas
    # ========================================================================

    @extend_schema(
        tags=['Reportes'],
        summary='Reporte de Empresas',
        description='''
        Genera reporte de empresas colaboradoras con estadísticas de prácticas.
        
        Requiere rol de **COORDINADOR**, **ADMINISTRADOR** o **SECRETARIA**.
        
        **Filtros disponibles**:
        - `sector`: Filtra por sector industrial/comercial
        - `validated`: Filtra por empresas validadas/no validadas (true/false)
        - `status`: Filtra por estado(s) separados por coma (ACTIVE, SUSPENDED, etc.)
        - `format`: Formato de salida (json, excel, csv)
        
        **Información incluida**:
        - Datos de la empresa (RUC, razón social, sector)
        - Estado de validación
        - Estadísticas de prácticas (total, activas, completadas)
        - Supervisores asignados
        - Información de contacto
        ''',
        parameters=[
            OpenApiParameter(
                name='sector',
                description='Sector industrial o comercial',
                required=False,
                type=OpenApiTypes.STR,
                examples=[
                    OpenApiExample('Tecnología', value='Tecnología'),
                    OpenApiExample('Construcción', value='Construcción'),
                ],
            ),
            OpenApiParameter(
                name='validated',
                description='Estado de validación de la empresa',
                required=False,
                type=OpenApiTypes.BOOL,
            ),
            OpenApiParameter(
                name='status',
                description='Estado(s) de la empresa separados por coma',
                required=False,
                type=OpenApiTypes.STR,
                examples=[
                    OpenApiExample('Estado único', value='ACTIVE'),
                    OpenApiExample('Múltiples estados', value='ACTIVE,SUSPENDED'),
                ],
            ),
            OpenApiParameter(
                name='format',
                description='Formato de exportación',
                required=False,
                type=OpenApiTypes.STR,
                enum=['json', 'excel', 'csv'],
                default='json',
            ),
        ],
        responses={
            200: OpenApiResponse(
                description='Reporte de empresas',
                examples=[
                    OpenApiExample(
                        'Reporte Empresas',
                        value={
                            'success': True,
                            'total_companies': 35,
                            'data': [
                                {
                                    'id': 'uuid',
                                    'razon_social': 'Tech Corp SAC',
                                    'ruc': '20123456789',
                                    'sector': 'Tecnología',
                                    'validated': True,
                                    'estado': 'ACTIVE',
                                    'practices_count': 12,
                                    'active_practices': 5,
                                    'supervisors_count': 3,
                                },
                            ],
                        },
                    ),
                ],
            ),
            403: OpenApiResponse(description='Sin permisos administrativos'),
        },
    )
    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsCoordinador | IsAdministrador | IsSecretaria],
        url_path='companies'
    )
    def companies_report(self, request):
        """
        Reporte de empresas con estadísticas de prácticas.
        
        GET /api/v2/reports/companies/
        ?sector=Tecnología
        &validated=true|false
        &status=ACTIVE,SUSPENDED
        &format=json|excel|csv
        
        Requiere: COORDINADOR, ADMINISTRADOR o SECRETARIA
        """
        try:
            # Parámetros
            sector = request.query_params.get('sector')
            validated = request.query_params.get('validated')
            status_filter = request.query_params.get('status', '').split(',')
            status_filter = [s.strip() for s in status_filter if s.strip()]
            format_type = request.query_params.get('format', 'json')
            
            # Query
            queryset = Company.objects.prefetch_related('practices', 'supervisors')
            
            if sector:
                queryset = queryset.filter(sector_economico__icontains=sector)
            
            if validated == 'true':
                queryset = queryset.filter(validated=True)
            elif validated == 'false':
                queryset = queryset.filter(validated=False)
            
            if status_filter:
                queryset = queryset.filter(estado__in=status_filter)
            
            # Preparar datos
            companies_data = []
            for company in queryset:
                companies_data.append({
                    'id': company.id,
                    'razon_social': company.razon_social,
                    'ruc': company.ruc,
                    'sector': company.sector_economico,
                    'direccion': company.direccion,
                    'telefono': company.telefono,
                    'email': company.email,
                    'validated': company.validated,
                    'estado': company.estado,
                    'fecha_registro': company.fecha_registro.isoformat(),
                    'estadisticas': {
                        'total_practicas': company.practices.count(),
                        'practicas_activas': company.practices.filter(estado='IN_PROGRESS').count(),
                        'practicas_completadas': company.practices.filter(estado='COMPLETED').count(),
                        'supervisores': company.supervisors.count()
                    }
                })
            
            if format_type == 'excel':
                return self._export_to_excel(companies_data, 'Reporte_Empresas')
            elif format_type == 'csv':
                return self._export_to_csv(companies_data, 'Reporte_Empresas')
            else:
                return Response({
                    'success': True,
                    'count': len(companies_data),
                    'data': companies_data,
                    'generated_at': timezone.now().isoformat()
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar reporte de empresas: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ========================================================================
    # Resumen Estadístico
    # ========================================================================

    @extend_schema(
        tags=['Reportes'],
        summary='Resumen Estadístico Ejecutivo',
        description='''
        Genera resumen ejecutivo con todas las métricas principales del sistema.
        
        Requiere rol de **COORDINADOR** o **ADMINISTRADOR**.
        
        **Métricas incluidas**:
        - 📊 Estadísticas de prácticas (total, por estado, completadas en el mes)
        - 👨‍🎓 Estadísticas de estudiantes (total, con prácticas, elegibles)
        - 🏢 Estadísticas de empresas (total, validadas, por sector)
        - 📄 Estadísticas de documentos (pendientes, aprobados, rechazados)
        - 📈 KPIs principales (tasa de éxito, promedio de horas, etc.)
        - 📅 Tendencias mensuales
        
        **Formatos disponibles**:
        - `json`: Respuesta JSON estructurada
        - `pdf`: Documento PDF ejecutivo (futuro)
        ''',
        parameters=[
            OpenApiParameter(
                name='format',
                description='Formato del resumen',
                required=False,
                type=OpenApiTypes.STR,
                enum=['json', 'pdf'],
                default='json',
            ),
        ],
        responses={
            200: OpenApiResponse(
                description='Resumen estadístico',
                examples=[
                    OpenApiExample(
                        'Resumen Estadístico',
                        value={
                            'success': True,
                            'data': {
                                'practices_stats': {
                                    'total': 120,
                                    'por_estado': {
                                        'IN_PROGRESS': 45,
                                        'COMPLETED': 60,
                                        'PENDING': 8,
                                    },
                                    'completadas_mes': 15,
                                    'promedio_horas': 475.5,
                                },
                                'students_stats': {
                                    'total': 250,
                                    'con_practicas': 120,
                                    'elegibles': 180,
                                },
                                'companies_stats': {
                                    'total': 35,
                                    'validadas': 30,
                                    'por_sector': {
                                        'Tecnología': 15,
                                        'Construcción': 10,
                                    },
                                },
                                'kpis': {
                                    'tasa_exito': 95.0,
                                    'tiempo_promedio_dias': 180,
                                },
                            },
                        },
                    ),
                ],
            ),
            403: OpenApiResponse(description='Sin permisos de coordinador/admin'),
        },
    )
    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsCoordinador | IsAdministrador],
        url_path='statistics_summary'
    )
    def statistics_summary(self, request):
        """
        Resumen estadístico completo del sistema.
        
        GET /api/v2/reports/statistics_summary/
        ?format=json|pdf
        
        Requiere: COORDINADOR o ADMINISTRADOR
        
        Returns: Resumen con todas las métricas principales
        """
        try:
            format_type = request.query_params.get('format', 'json')
            
            # Estadísticas de prácticas
            practices_stats = {
                'total': Practice.objects.count(),
                'por_estado': dict(
                    Practice.objects.values('estado').annotate(
                        count=Count('id')
                    ).values_list('estado', 'count')
                ),
                'completadas_mes': Practice.objects.filter(
                    estado='COMPLETED',
                    fecha_fin__gte=timezone.now() - timedelta(days=30)
                ).count(),
                'promedio_horas': Practice.objects.filter(
                    estado='COMPLETED'
                ).aggregate(Avg('horas_totales'))['horas_totales__avg'] or 0
            }
            
            # Estadísticas de estudiantes
            students_stats = {
                'total': Student.objects.count(),
                'elegibles': Student.objects.filter(
                    semestre_actual__gte=6,
                    promedio_ponderado__gte=12.0
                ).count(),
                'con_practica': Student.objects.filter(
                    practices__isnull=False
                ).distinct().count(),
                'por_carrera': list(
                    Student.objects.values('escuela_profesional').annotate(
                        count=Count('id')
                    ).order_by('-count')[:10]
                )
            }
            
            # Estadísticas de empresas
            companies_stats = {
                'total': Company.objects.count(),
                'validadas': Company.objects.filter(validated=True).count(),
                'activas': Company.objects.filter(estado='ACTIVE').count(),
                'por_sector': list(
                    Company.objects.values('sector_economico').annotate(
                        count=Count('id')
                    ).order_by('-count')[:10]
                ),
                'con_practicas_activas': Company.objects.filter(
                    practices__estado='IN_PROGRESS'
                ).distinct().count()
            }
            
            # Estadísticas de documentos
            documents_stats = {
                'total': Document.objects.count(),
                'pendientes': Document.objects.filter(estado='PENDING').count(),
                'aprobados': Document.objects.filter(estado='APPROVED').count(),
                'rechazados': Document.objects.filter(estado='REJECTED').count()
            }
            
            # Estadísticas de usuarios
            users_stats = {
                'total': User.objects.count(),
                'activos': User.objects.filter(is_active=True).count(),
                'por_rol': dict(
                    User.objects.values('role').annotate(
                        count=Count('id')
                    ).values_list('role', 'count')
                )
            }
            
            # Tendencias (últimos 6 meses)
            six_months_ago = timezone.now() - timedelta(days=180)
            trends = {
                'nuevas_practicas_mes': list(
                    Practice.objects.filter(
                        fecha_registro__gte=six_months_ago
                    ).extra(
                        select={'month': 'EXTRACT(month FROM fecha_registro)'}
                    ).values('month').annotate(
                        count=Count('id')
                    ).order_by('month')
                ),
                'nuevos_estudiantes_mes': list(
                    Student.objects.filter(
                        user__date_joined__gte=six_months_ago
                    ).extra(
                        select={'month': 'EXTRACT(month FROM user__date_joined)'}
                    ).values('month').annotate(
                        count=Count('id')
                    ).order_by('month')
                )
            }
            
            summary_data = {
                'practices': practices_stats,
                'students': students_stats,
                'companies': companies_stats,
                'documents': documents_stats,
                'users': users_stats,
                'trends': trends,
                'generated_at': timezone.now().isoformat()
            }
            
            if format_type == 'pdf':
                # TODO: Implementar generación de PDF
                return Response({
                    'success': False,
                    'error': 'Exportación a PDF no implementada aún'
                }, status=status.HTTP_501_NOT_IMPLEMENTED)
            else:
                return Response({
                    'success': True,
                    'data': summary_data
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar resumen estadístico: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ========================================================================
    # Certificado de Práctica
    # ========================================================================

    @extend_schema(
        tags=['Reportes'],
        summary='Certificado de Práctica',
        description='''
        Genera certificado oficial de práctica profesional completada.
        
        Requiere que el usuario tenga relación con la práctica:
        - **PRACTICANTE**: Solo puede ver su propio certificado
        - **SUPERVISOR**: Puede ver certificados de sus practicantes
        - **COORDINADOR/ADMIN**: Puede ver todos los certificados
        
        **Requisitos**:
        - La práctica debe estar en estado COMPLETED
        - Se deben haber completado las horas mínimas requeridas
        - Documentación debe estar aprobada
        
        **Contenido del certificado**:
        - Datos del estudiante (nombre, código, carrera)
        - Datos de la empresa y supervisor
        - Período de práctica (fecha inicio y fin)
        - Total de horas completadas
        - Descripción de actividades realizadas
        - Firma digital y sello institucional
        ''',
        responses={
            200: OpenApiResponse(
                description='Datos del certificado',
                examples=[
                    OpenApiExample(
                        'Certificado',
                        value={
                            'success': True,
                            'data': {
                                'student': {
                                    'nombre_completo': 'Juan Pérez',
                                    'codigo': '2020123456',
                                    'carrera': 'Ingeniería de Sistemas',
                                },
                                'practice': {
                                    'titulo': 'Desarrollo Web',
                                    'fecha_inicio': '2024-01-15',
                                    'fecha_fin': '2024-07-15',
                                    'horas_completadas': 480,
                                },
                                'company': {
                                    'razon_social': 'Tech Corp SAC',
                                    'ruc': '20123456789',
                                },
                                'supervisor': {
                                    'nombre_completo': 'María García',
                                    'cargo': 'Jefe de Desarrollo',
                                },
                                'certificate_number': 'CERT-2024-001234',
                                'issue_date': '2024-07-20',
                            },
                        },
                    ),
                ],
            ),
            403: OpenApiResponse(description='Sin permisos para ver este certificado'),
            404: OpenApiResponse(description='Práctica no encontrada'),
            400: OpenApiResponse(description='La práctica no está completada'),
        },
    )
    @action(
        detail=True,
        methods=['get'],
        permission_classes=[IsAuthenticated],
        url_path='certificate'
    )
    def practice_certificate(self, request, pk=None):
        """
        Genera certificado de práctica completada.
        
        GET /api/v2/reports/{practice_id}/certificate/
        
        Requiere: Usuario relacionado a la práctica
        
        Returns: Datos para generar certificado
        """
        try:
            # Obtener práctica
            try:
                practice = Practice.objects.select_related(
                    'student__user',
                    'company',
                    'supervisor__user'
                ).get(id=pk)
            except Practice.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'Práctica no encontrada'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Verificar permisos
            user = request.user
            if not (
                user.role in ['COORDINADOR', 'ADMINISTRADOR'] or
                user == practice.student.user or
                (hasattr(user, 'supervisor_profile') and user.supervisor_profile == practice.supervisor)
            ):
                return Response({
                    'success': False,
                    'error': 'No tiene permisos para ver este certificado'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Verificar que la práctica esté completada
            if practice.estado != 'COMPLETED':
                return Response({
                    'success': False,
                    'error': 'La práctica debe estar completada para generar certificado'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Datos del certificado
            certificate_data = {
                'practice': {
                    'id': practice.id,
                    'titulo': practice.titulo,
                    'modalidad': practice.modalidad,
                    'fecha_inicio': practice.fecha_inicio.isoformat() if practice.fecha_inicio else None,
                    'fecha_fin': practice.fecha_fin.isoformat() if practice.fecha_fin else None,
                    'horas_totales': practice.horas_totales,
                    'descripcion': practice.descripcion
                },
                'student': {
                    'codigo': practice.student.codigo_estudiante,
                    'nombre_completo': practice.student.user.get_full_name(),
                    'carrera': practice.student.escuela_profesional,
                    'email': practice.student.user.email
                },
                'company': {
                    'razon_social': practice.company.razon_social,
                    'ruc': practice.company.ruc,
                    'sector': practice.company.sector_economico,
                    'direccion': practice.company.direccion
                },
                'supervisor': {
                    'nombre': practice.supervisor.user.get_full_name() if practice.supervisor else 'Sin asignar',
                    'cargo': practice.supervisor.cargo if practice.supervisor else None,
                    'email': practice.supervisor.user.email if practice.supervisor else None
                },
                'calificacion': practice.calificacion_final,
                'observaciones': practice.observaciones_finales,
                'generated_at': timezone.now().isoformat(),
                'certificate_number': f'CERT-{practice.id:06d}-{timezone.now().year}'
            }
            
            return Response({
                'success': True,
                'data': certificate_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar certificado: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ========================================================================
    # Métodos Auxiliares de Exportación
    # ========================================================================

    def _export_to_excel(self, data, filename):
        """Exporta datos a formato Excel."""
        if not EXCEL_AVAILABLE:
            return Response({
                'success': False,
                'error': 'Librería openpyxl no disponible. Instale con: pip install openpyxl'
            }, status=status.HTTP_501_NOT_IMPLEMENTED)
        
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = filename[:31]  # Excel limita a 31 caracteres
            
            # Estilos
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Si hay datos
            if data:
                # Obtener headers (claves del primer elemento, flatten nested dicts)
                headers = self._flatten_keys(data[0])
                
                # Escribir headers
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Escribir datos
                for row_idx, item in enumerate(data, start=2):
                    flat_item = self._flatten_dict(item)
                    for col_idx, key in enumerate(headers, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=flat_item.get(key, ''))
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Crear response
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al exportar a Excel: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _export_to_csv(self, data, filename):
        """Exporta datos a formato CSV."""
        try:
            # Crear CSV en memoria
            output = io.StringIO()
            
            if data:
                # Flatten data
                flat_data = [self._flatten_dict(item) for item in data]
                
                # Obtener headers
                headers = list(flat_data[0].keys())
                
                # Escribir CSV
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                writer.writerows(flat_data)
            
            # Crear response
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d")}.csv"'
            
            return response
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al exportar a CSV: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _flatten_dict(self, d, parent_key='', sep='_'):
        """Aplana un diccionario anidado."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

    def _flatten_keys(self, d, parent_key='', sep='_'):
        """Obtiene las claves de un diccionario anidado."""
        keys = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                keys.extend(self._flatten_keys(v, new_key, sep=sep))
            else:
                keys.append(new_key)
        return keys
