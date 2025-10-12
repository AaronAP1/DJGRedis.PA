import io
from typing import List, Dict, Any

from django.http import HttpResponse
from django.db import transaction
from django.conf import settings
from django.core.files.uploadedfile import UploadedFile

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from drf_spectacular.utils import extend_schema, OpenApiResponse, OpenApiParameter

from openpyxl import Workbook, load_workbook

from src.adapters.secondary.database.models import User, Student
from src.infrastructure.security.permissions import IsAdminOnly
from .serializers import (
    ImportPreviewRowSerializer,
    ImportConfirmResponseSerializer,
)


ALL_COLUMNS = ['codigo', 'nombre_completo', 'apellido_completo', 'email', 'role', 'estado']
ALLOWED_ROLES = {'PRACTICANTE'}  # Solo practicantes en importación
ALLOWED_EMAIL_DOMAINS = {d.lower() for d in getattr(settings, 'ALLOWED_EMAIL_DOMAINS', [])}


def _build_username(nombre_completo: str, apellido_completo: str) -> str:
    def _first_token(s: str) -> str:
        return (s.strip().split()[0] if s and s.strip() else '')
    n = _first_token(nombre_completo).lower()
    a = _first_token(apellido_completo).lower()
    base = '.'.join(filter(None, [n, a])) or 'usuario'
    return base


class ImportTemplateXLSXView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        operation_id='users_import_template_xlsx',
        tags=['Users Import'],
        summary='Descargar plantilla Excel para importación de usuarios PRACTICANTE',
        description='Descarga una plantilla Excel con el formato correcto para importar usuarios con rol PRACTICANTE. Incluye código estudiantil y campos requeridos.',
        responses={200: OpenApiResponse(description='XLSX template for PRACTICANTE users')},
    )
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Usuarios'
        ws.append(ALL_COLUMNS)
        ws.append([
            '2021000001', 'Cristian Antony', 'Lara Arcos', 'cristian.lara@upeu.edu.pe', 'PRACTICANTE', 'ACTIVO'
        ])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="users_template.xlsx"'
        return resp


def _read_rows_from_file(upload: UploadedFile) -> List[Dict[str, Any]]:
    name = (upload.name or '').lower()
    if name.endswith('.xlsx'):
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else '' for c in next(rows_iter)]
        result = []
        for row in rows_iter:
            obj = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers)) if headers[i]}
            result.append(obj)
        return result
    raise ValueError('Formato no soportado. Debe ser .xlsx')


def _validate_row(idx: int, row: Dict[str, Any], seen_emails: set) -> Dict[str, Any]:
    errors: List[str] = []
    email = (str(row.get('email') or '')).strip()
    nombre = (str(row.get('nombre_completo') or '')).strip()
    apellidos = (str(row.get('apellido_completo') or '')).strip()
    codigo = (str(row.get('codigo') or '')).strip()
    role = (str(row.get('role') or '')).strip().upper()
    estado = (str(row.get('estado') or '')).strip().upper()

    if role not in ALLOWED_ROLES:
        errors.append('solo se permite importar PRACTICANTE')

    if not email:
        errors.append('email requerido')
    if email and ALLOWED_EMAIL_DOMAINS:
        try:
            domain = email.split('@')[1].lower()
            if domain not in ALLOWED_EMAIL_DOMAINS:
                errors.append('email dominio no permitido')
        except Exception:
            errors.append('email inválido')

    if email in seen_emails:
        errors.append('email duplicado en archivo')

    if not nombre:
        errors.append('nombre_completo requerido')
    if not apellidos:
        errors.append('apellido_completo requerido')
    if not codigo:
        errors.append('codigo requerido')

    if email and User.objects.filter(email__iexact=email).exists():
        errors.append('email ya registrado')

    status = 'invalid' if errors else 'valid'
    seen_emails.add(email)

    return {
        'index': idx,
        'email': email,
        'first_name': nombre,
        'last_name': apellidos,
        'role': role,
        'status': status,
        'errors': errors,
        'codigo': codigo,
        'estado': estado,
    }



class ImportConfirmView(APIView):
    # Solo ADMINISTRADOR puede ejecutar la creación real de usuarios
    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        operation_id='users_import_confirm',
        tags=['Users Import'],
        summary='Importar usuarios PRACTICANTE masivamente (Excel)',
        description='Importa usuarios con rol PRACTICANTE desde archivo Excel. Username generado automáticamente por nombres y apellidos. La contraseña se establece como el código de estudiante. Envía email de bienvenida automáticamente.',
        request={'multipart/form-data': {'type': 'object', 'properties': {
            'file': {'type': 'string', 'format': 'binary', 'description': 'Archivo Excel con usuarios PRACTICANTE'},
            'send_email': {'type': 'boolean', 'description': 'Enviar email de bienvenida (default: true)'}
        }}},
        responses={200: ImportConfirmResponseSerializer},
    )
    def post(self, request):
        upload = request.FILES.get('file')
        send_email = str(request.data.get('send_email', 'true')).lower() in ('1', 'true', 'yes')
        if not upload:
            return Response({'detail': 'Falta archivo (file)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = _read_rows_from_file(upload)
        except Exception as e:
            return Response({'detail': f'Archivo inválido: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        seen = set()
        preview_rows = []
        for i, row in enumerate(rows, start=2):
            normalized = {k.strip(): ('' if v is None else str(v).strip()) for k, v in row.items()}
            pr = _validate_row(i, normalized, seen)
            preview_rows.append(pr)

        valid_rows = [r for r in preview_rows if r['status'] == 'valid']
        created_count = 0

        from src.infrastructure.security.tasks import send_welcome_email

        with transaction.atomic():
            for r in valid_rows:
                email = r['email']
                role = r['role']
                first_name = r['first_name']
                last_name = r['last_name']
                codigo = r.get('codigo') or ''
                estado = (r.get('estado') or '').upper()
                is_active = True if estado in ('ACTIVO', 'ACTIVE', '1', 'TRUE', 'SI', 'YES') else False

                # Generar username según regla: primer nombre + '.' + primer apellido (lowercase)
                username = _build_username(first_name, last_name)

                user = User.objects.create_user(
                    email=email,
                    password=codigo,
                    first_name=first_name,
                    last_name=last_name,
                    role='PRACTICANTE',
                    is_active=is_active,
                    username=username,
                )
                created_count += 1

                Student.objects.create(
                    user=user,
                    codigo_estudiante=codigo,
                )

                # Enviar email de bienvenida (COMENTADO - servidor de correo no configurado)
                # if send_email and getattr(settings, 'EMAIL_ENABLED', False):
                #     try:
                #         send_welcome_email.delay(str(user.id))
                #     except Exception:
                #         # fallback si Celery no está corriendo
                #         from django.core.mail import EmailMultiAlternatives
                #         from django.template.loader import render_to_string
                #         ctx = {
                #             'user': user, 
                #             'frontend_url': getattr(settings, 'FRONTEND_URL', '')
                #         }
                #         html_body = render_to_string('emails/user_welcome.html', ctx)
                #         msg = EmailMultiAlternatives(
                #             'Bienvenido al Sistema de Prácticas', 
                #             html_body, 
                #             settings.DEFAULT_FROM_EMAIL, 
                #             [email]
                #         )
                #         msg.attach_alternative(html_body, 'text/html')
                #         msg.send()

        valid_count = len(valid_rows)
        invalid_count = len(preview_rows) - valid_count
        data = {
            'created_count': created_count,
            'rows': preview_rows,
            'valid_count': valid_count,
            'invalid_count': invalid_count,
        }
        return Response(data, status=status.HTTP_200_OK)
